import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime

class TeacherDashboard:
    def __init__(self, root):
        self.root = root
        self.root.title("Teacher Dashboard")
        self.file_name = "teacher_data.xlsx"
        self.setup_main_window()

    # Function to set up the main window with buttons for different actions
    def setup_main_window(self):
        tk.Button(self.root, text="Add Teacher", command=self.open_add_teacher_window).grid(row=0, column=0, padx=10, pady=10)
        tk.Button(self.root, text="Add New Batch", command=self.open_add_batch_window).grid(row=1, column=0, padx=10, pady=10)
        tk.Button(self.root, text="Calculate Salary", command=self.open_salary_calculation_window).grid(row=2, column=0, padx=10, pady=10)

    # Function to open the window for adding a new teacher
    def open_add_teacher_window(self):
        self.add_teacher_window = tk.Toplevel(self.root)
        self.add_teacher_window.title("Add Teacher")

        tk.Label(self.add_teacher_window, text="Teacher Name:").grid(row=0, column=0, padx=10, pady=10)
        self.entry_teacher_name = tk.Entry(self.add_teacher_window, width=30)
        self.entry_teacher_name.grid(row=0, column=1, padx=10, pady=10)

        tk.Label(self.add_teacher_window, text="Level:").grid(row=1, column=0, padx=10, pady=10)
        self.entry_level = tk.Entry(self.add_teacher_window, width=30)
        self.entry_level.grid(row=1, column=1, padx=10, pady=10)

        tk.Label(self.add_teacher_window, text="Remuneration Slabs \n(Minimum Maximum Remuneration):").grid(row=2, column=0, padx=10, pady=10)
        self.text_remuneration_slabs = tk.Text(self.add_teacher_window, height=5, width=30)
        self.text_remuneration_slabs.grid(row=2, column=1, padx=10, pady=10)

        tk.Button(self.add_teacher_window, text="Submit", command=self.submit_teacher_form).grid(row=3, column=0, columnspan=2, pady=10)

    # Function to open the window for adding a new batch
    def open_add_batch_window(self):
        self.add_batch_window = tk.Toplevel(self.root)
        self.add_batch_window.title("Add New Batch")

        tk.Label(self.add_batch_window, text="Teacher Name:").grid(row=0, column=0, padx=10, pady=10)
        self.entry_batch_teacher_name = ttk.Combobox(self.add_batch_window, width=28)
        self.entry_batch_teacher_name.grid(row=0, column=1, padx=10, pady=10)
        self.entry_batch_teacher_name.bind("<Return>", lambda event: self.suggest_teachers(event, "Salary Slab"))
        self.entry_batch_teacher_name.bind("<<ComboboxSelected>>", self.load_teacher_levels)

        tk.Label(self.add_batch_window, text="Level:").grid(row=1, column=0, padx=10, pady=10)
        self.combo_batch_level = ttk.Combobox(self.add_batch_window, width=28)
        self.combo_batch_level.grid(row=1, column=1, padx=10, pady=10)

        tk.Label(self.add_batch_window, text="Start Date (DD-MM-YYYY):").grid(row=2, column=0, padx=10, pady=10)
        self.entry_start_date = tk.Entry(self.add_batch_window, width=30)
        self.entry_start_date.grid(row=2, column=1, padx=10, pady=10)

        tk.Label(self.add_batch_window, text="Batch Timing:").grid(row=3, column=0, padx=10, pady=10)
        self.entry_batch_time = tk.Entry(self.add_batch_window, width=30)
        self.entry_batch_time.grid(row=3, column=1, padx=10, pady=10)

        tk.Label(self.add_batch_window, text="Duration:").grid(row=4, column=0, padx=10, pady=10)
        self.entry_duration = tk.Entry(self.add_batch_window, width=30)
        self.entry_duration.grid(row=4, column=1, padx=10, pady=10)

        tk.Label(self.add_batch_window, text="Number of Students:").grid(row=5, column=0, padx=10, pady=10)
        self.entry_num_students = tk.Entry(self.add_batch_window, width=30)
        self.entry_num_students.grid(row=5, column=1, padx=10, pady=10)

        tk.Button(self.add_batch_window, text="Submit", command=self.submit_batch_form).grid(row=6, column=0, columnspan=2, pady=10)

    # Function to open the window for salary calculation
    def open_salary_calculation_window(self):
        self.salary_window = tk.Toplevel(self.root)
        self.salary_window.title("Calculate Salary")

        tk.Label(self.salary_window, text="Teacher Name:").grid(row=0, column=0, padx=10, pady=10)
        self.entry_batch_teacher_name = ttk.Combobox(self.salary_window, width=28)
        self.entry_batch_teacher_name.grid(row=0, column=1, padx=10, pady=10)
        self.entry_batch_teacher_name.bind("<Return>", lambda event: self.suggest_teachers(event, "Batches"))
        self.entry_batch_teacher_name.bind("<<ComboboxSelected>>", self.show_batches)

        tk.Label(self.salary_window, text="Batch Code:").grid(row=1, column=0, padx=10, pady=10)
        self.combo_salary_batch_code = ttk.Combobox(self.salary_window, width=28)
        self.combo_salary_batch_code.grid(row=1, column=1, padx=10, pady=10)
        self.combo_salary_batch_code.bind("<<ComboboxSelected>>", self.load_batch_details)

        tk.Label(self.salary_window, text="Number of Hours:").grid(row=4, column=0, padx=10, pady=10)
        self.entry_num_hours = tk.Entry(self.salary_window, width=30)
        self.entry_num_hours.grid(row=4, column=1, padx=10, pady=10)

        tk.Button(self.salary_window, text="Calculate Salary", command=self.calculate_salary).grid(row=5, column=0, columnspan=2, pady=10)

    # Function to save data to an Excel file
    def save_to_excel(self, data, sheet_name, headers):
        try:
            wb = load_workbook(self.file_name)
        except FileNotFoundError:
            wb = Workbook()

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
        else:
            ws = wb[sheet_name]

        for entry in data:
            ws.append(entry)

        wb.save(self.file_name)
        
    # Function to handle the submission of the teacher form
    def submit_teacher_form(self):
        teacher_name = self.entry_teacher_name.get()
        level = self.entry_level.get()
        slab_data = self.text_remuneration_slabs.get("1.0", tk.END).strip().split("\n")
        data = []

        for slab in slab_data:
            try:
                min_students, max_students, remuneration = map(int, slab.split())
                data.append([teacher_name, level, min_students, max_students, remuneration])
            except ValueError:
                messagebox.showerror("Input Error", "Please enter valid data for remuneration slabs (Minimum Maximum Remuneration)")
                return

        self.save_to_excel(data, "Salary Slab", ["Teacher Name", "Level", "Min Students", "Max Students", "Remuneration"])
        messagebox.showinfo("Success", "Data saved successfully")
        self.entry_teacher_name.delete(0, tk.END)
        self.entry_level.delete(0, tk.END)
        self.text_remuneration_slabs.delete("1.0", tk.END)

    # Function to handle the submission of the batch form
    def submit_batch_form(self):
        batch_time = self.entry_batch_time.get()
        batch_level = self.combo_batch_level.get()
        teacher_name = self.entry_batch_teacher_name.get()
        start_date = self.entry_start_date.get()
        duration = self.entry_duration.get()
        num_students = self.entry_num_students.get()
        
        try:
            start_date = datetime.strptime(start_date, '%d-%m-%Y')
        except ValueError:
            messagebox.showerror("Date Error", "Please enter dates in DD-MM-YYYY format")
            return
        
        try:
            duration = int(duration)
            num_students = int(num_students)
        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numbers for duration and number of students")
            return

        batch_code = f"{batch_level}{batch_time}-{teacher_name[:1].upper()}{teacher_name.split(' ')[1][:1].upper()}-{start_date.strftime('%d%m%Y')}"
        data = [[teacher_name, batch_level, batch_code, start_date.strftime('%d-%m-%Y'), batch_time, duration, num_students, "Ongoing"]]
        self.save_to_excel(data, "Batches", ["Teacher Name", "Batch Level", "Batch Code", "Start Date","Batch Time", "Duration", "Number of Students", "Status"])
        
        
        # Adding data to Report sheet
        data = [[batch_code, teacher_name, start_date.strftime('%d-%m-%Y'), duration, datetime.now().strftime('%d-%m-%Y'), duration]]
        self.save_to_excel(data, "Report", ["Batch Code", "Teacher Name", "Start Date", "Duration", "Updated On", "Remaining Duration"])
        
        
        # Append batch details to teacher's individual sheet
        try:
            wb = load_workbook(self.file_name)
            if teacher_name not in wb.sheetnames:
                wb.create_sheet(title=teacher_name)  # Create sheet if it doesn't exist               
            ws_teacher = wb[teacher_name]
            # Add the default Cumulative columns
            ws_teacher.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws_teacher.cell(row=1, column=1, value="Cumulative")
            ws_teacher.cell(row=2, column=1, value="Date")
            ws_teacher.cell(row=2, column=2, value="Total")
            
        except FileNotFoundError:
            messagebox.showerror("Error", "Teacher data file not found")
            return

        # Find the next available row in the teacher's sheet
        max_col = ws_teacher.max_column+1
        # Write batch data horizontally
        ws_teacher.merge_cells(start_row=1, start_column=max_col+1, end_row=1, end_column=max_col+4)
        ws_teacher.cell(row=1, column=max_col+1, value=batch_code)
        ws_teacher.cell(row=2, column=max_col+1, value="Date")
        ws_teacher.cell(row=2, column=max_col+2, value="Students (Remuneration)")
        ws_teacher.cell(row=2, column=max_col+3, value="Duration")
        ws_teacher.cell(row=2, column=max_col+4, value="Salary")
        
        wb.save(self.file_name)  
        messagebox.showinfo("Success", "Batch data saved successfully")
        
        # Clear input fields after submission
        self.combo_batch_level.delete(0, tk.END)
        self.entry_batch_time.delete(0, tk.END)
        self.entry_start_date.delete(0, tk.END)
        self.entry_duration.delete(0, tk.END)
        self.entry_num_students.delete(0, tk.END)


    
    # Function to suggest teacher names as the user types - used for Adding Batch and Calculating Salary
    def suggest_teachers(self, event, sheet_name):
        typed_text = self.entry_batch_teacher_name.get().lower()
        try:
            wb = load_workbook(self.file_name)
            ws = wb[sheet_name]
        except FileNotFoundError:
            messagebox.showerror("Error", "Teacher data file not found")
            return

        teacher_names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if typed_text in row[0].lower():
                teacher_names.add(row[0])

        self.entry_batch_teacher_name['values'] = list(teacher_names)
        self.entry_batch_teacher_name.event_generate('<Down>')
        

    def load_teacher_levels(self, event):
        teacher_name = self.entry_batch_teacher_name.get()
        if not teacher_name:
            return

        try:
            wb = load_workbook(self.file_name)
            ws = wb['Salary Slab']
        except FileNotFoundError:
            messagebox.showerror("Error", "Teacher data file not found")
            return

        levels = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == teacher_name:
                levels.add(row[1])

        self.combo_batch_level['values'] = list(levels)
        if levels:
            self.combo_batch_level.current(0)


    def show_batches(self, event):
        teacher_name = self.entry_batch_teacher_name.get()
        if not teacher_name:
            return

        try:
            wb = load_workbook(self.file_name)
            ws = wb['Batches']
        except FileNotFoundError:
            messagebox.showerror("Error", "Teacher data file not found")
            return

        batches = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == teacher_name and row[7] == "Ongoing":
                batches.add(row[2])

        self.combo_salary_batch_code['values'] = list(batches)
        if batches:
            self.combo_salary_batch_code.current(0)


    # Function to load batch details and number of students for the selected batch - used for Calculating Salary
    def load_batch_details(self, event):
        batch_code = self.combo_salary_batch_code.get()
        teacher_name = self.entry_batch_teacher_name.get()

        try:
            wb = load_workbook(self.file_name)
            ws = wb['Batches']
        except FileNotFoundError:
            messagebox.showerror("Error", "Teacher data file not found")
            return

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == teacher_name and row[2] == batch_code:
                self.num_students = row[4]
                break


    def calculate_salary(self):
        teacher_name = self.entry_batch_teacher_name.get()
        batch_code = self.combo_salary_batch_code.get()
        num_hours = int(self.entry_num_hours.get())

        try:
            wb = load_workbook(self.file_name)
            ws_batches = wb['Batches']
            ws_salary_slab = wb['Salary Slab']
            ws_report = wb['Report']
            ws_teacher = wb[teacher_name]
        except FileNotFoundError:
            messagebox.showerror("Error", "Teacher data file not found")
            return

        # Find batch details
        batch_data = None
        for row in ws_batches.iter_rows(min_row=2, values_only=True):
            if row[0] == teacher_name and row[2] == batch_code:
                batch_data = row
                break
        if not batch_data:
            messagebox.showerror("Error", "Batch data not found")
            return

        # Find remuneration from salary slab
        remuneration = None
        for row in ws_salary_slab.iter_rows(min_row=2, values_only=True):
            if row[0] == teacher_name and row[2] <= batch_data[6] <= row[3]:  
                remuneration = row[4]
                break
        if remuneration is None:
            messagebox.showerror("Error", "No matching salary slab found")
            return
            
        

        # Calculate salary
        salary = remuneration * num_hours

        # Find the column index of the batch code
        batch_col = None
        for col in range(1, ws_teacher.max_column + 1):
            if ws_teacher.cell(row=1, column=col).value == batch_code:
                batch_col = col
                break
        if not batch_col:
            messagebox.showerror("Error", "Batch code not found in teacher's sheet")
            return

        # Find or create a row for the current month
        current_month = datetime.now().strftime('%m-%Y')
        empty_row = None
        for row in range(3, ws_teacher.max_row + 1):
            cell_value = ws_teacher.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str) and current_month in cell_value:
                empty_row = row
                break

        # If no row for the current month is found, create a new row
        if empty_row is None:
            empty_row = ws_teacher.max_row + 1

        # Write batch data horizontally
        ws_teacher.cell(row=empty_row, column=batch_col, value=f"{datetime.now().strftime('%d-%m-%Y')}")
        ws_teacher.cell(row=empty_row, column=batch_col + 1, value=f'{batch_data[6]}({remuneration})')
        ws_teacher.cell(row=empty_row, column=batch_col + 2, value=num_hours)
        ws_teacher.cell(row=empty_row, column=batch_col + 3, value=salary)

        # Update cumulative salary
        cumulative_col = 1  # Assuming cumulative salary is in the 2nd column
        cumulative_salary = int(ws_teacher.cell(row=empty_row, column=cumulative_col + 1).value or 0)
        ws_teacher.cell(row=empty_row, column=cumulative_col, value=f"{datetime.now().strftime('%d-%m-%Y')}")
        ws_teacher.cell(row=empty_row, column=cumulative_col + 1, value=cumulative_salary + salary)
        
        
        # Update the hours in the "Report" sheet
        for row in ws_report.iter_rows(min_row=2, values_only=False):  # min_row=2 to skip header
            if row[0].value ==  batch_code and row[1].value == teacher_name:
                row[4].value = datetime.now().strftime('%d-%m-%Y')
                row[5].value -= num_hours

                if row[5].value <= 0:
                    for batch_row in ws_batches.iter_rows(min_row=2, values_only=False):
                        if batch_row[0].value == teacher_name and batch_row[2].value == batch_code:
                            batch_row[7].value = "Completed"
                            break
                break

        wb.save(self.file_name)

        # Clear input fields after submission
        self.combo_salary_batch_code.delete(0, tk.END)
        self.entry_num_hours.delete(0, tk.END)

if __name__ == "__main__":
    root = tk.Tk()
    app = TeacherDashboard(root)
    root.mainloop()
