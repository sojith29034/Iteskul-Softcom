import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import requests


def run_main_app():      
    st.set_page_config(page_title="Student Attendance Report", layout="wide")
    
    st.markdown("""
        <style>
            .reportview-container {margin-top: -2em;}
            .st-emotion-cache-1jicfl2 {padding: 2rem 3rem 10rem;}
            h1#student-attendance-report {text-align: center;}
            header #MainMenu {visibility: hidden; display: none;}
            .stActionButton {visibility: hidden; display: none;}
            .stDeployButton {display:none;}
            footer {visibility: hidden;}
            stDecoration {display:none;}
            .stTabs button {margin-right: 50px;}
            .viewerBadge_container__r5tak {display: none;}
            p.credits {user-select: none; filter: opacity(0);}
        </style>
    """, unsafe_allow_html=True)
    
    
    # Function to fetch sample files from GitHub (adjust the URL and filenames as needed)
    def fetch_sample_files():
        base_url = 'https://github.com/sojith29034/Iteskul-Softcom/raw/main_branch/StudentData/'
        files = [
            'German A1-WD-08.00pm-VP-23052024.xlsx',
            'Japanese N3-TR-2.00pm -SS-240923.xlsx',
            'Japanese N5-WN-2.00pm -WT-240923.xlsx'
        ]
        uploaded_files = []
    
        for file in files:
            url = base_url + file
            response = requests.get(url)
            if response.status_code == 200:
                # Append tuple with name and BytesIO content
                uploaded_files.append((file, BytesIO(response.content)))
            else:
                st.warning(f"Failed to fetch {file} from GitHub.")
    
        return uploaded_files

        
    
    # Helper function to calculate attendance percentage
    def calculate_attendance(df):
        # Calculate total sessions and present sessions
        df['Total Sessions'] = df.iloc[:, 1:].apply(lambda row: row.count(), axis=1)
        df['Present Sessions'] = df.iloc[:, 1:].apply(lambda row: row.value_counts().get('P', 0) + row.value_counts().get('p', 0), axis=1)
        df['Attendance %'] = (df['Present Sessions'] / df['Total Sessions']) * 100
    
        # Reorder columns to have Student Name, Total Sessions, Present Sessions, and Attendance % at the start
        columns_order = ['Student Name', 'Total Sessions', 'Present Sessions', 'Attendance %'] + [col for col in df.columns if col not in ['Student Name', 'Total Sessions', 'Present Sessions', 'Attendance %']]
        df = df[columns_order]
    
        return df
    
    # Helper function to find students who left
    def students_left(df):
        left = []
        for index, row in df.iterrows():
            attendance = row[:-3]  # Get the last three attendance columns before the summary columns
            attendance = attendance.str.upper()  # Convert to uppercase 
            
            # Check if any student has "Left"
            if attendance.str.contains("LEFT").any():
                left.append(row['Student Name'])
        
        return left
    
    
    def find_consecutive_absentees(df):
        absentees = []
        for index, row in df.iterrows():
            attendance_series = row[-6:-3]  # Get the last three attendance columns before the summary columns
            attendance_series = attendance_series.str.upper()  # Convert to uppercase to handle both 'A' and 'a'
            
            # Check if all three values in the attendance_series are 'A'
            if attendance_series.tolist() == ['A', 'A', 'A']:
                absentees.append(row['Student Name'])
        
        return absentees
    
    # Helper function to find students absent for a total of 5 or more days
    def find_absentees(df):
        five_absentees = []
        ten_absentees = []
        
        for index, row in df.iterrows():
            attendance_series = row[3:]  # Get all attendance columns after the summary columns
            attendance_series = attendance_series.str.upper()  # Convert to uppercase to handle both 'A' and 'a'
            
            # Count total absences
            total_absences = attendance_series.tolist().count('A')
            
            if row['Total Sessions'] >= 25:
                if total_absences >= 10:
                    ten_absentees.append(row['Student Name'])
                    five_absentees.append(row['Student Name'])
            if row['Total Sessions'] >= 10:
                if total_absences >= 5:
                    if row['Student Name'] not in five_absentees:
                        five_absentees.append(row['Student Name']) 
        
        # Return absentees or default statement
        if ten_absentees:
            return five_absentees, ten_absentees
        elif five_absentees:
            return five_absentees, []
        else:
            return [], []
        
    # Highlight function
    def highlight_rows(row, condition_list):
        if row['Student Name'] in condition_list:
            return ['background-color: yellow'] * len(row)
        return [''] * len(row)
    
    
  
    # Helper function to pad lists to the same length
    def pad_lists(lists):
        max_len = max(len(lst) for lst in lists)
        return [lst + [""] * (max_len - len(lst)) for lst in lists]
    
    # Function to generate excel sheets
    def generate_excel(reports):
        output = BytesIO()
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            trainer_report_data = []
    
            for class_name, report in reports.items():
                # Add class attendance data to a sheet
                df = report["Attendance Data"]
                df.to_excel(writer, sheet_name=f"{class_name[:20]} Data", index=False, startrow=1)
    
                # Add class summary to a separate sheet
                summary_data = {
                    "Attendance < 75%": report["Low Attendance"],
                    "3 Consecutive Absents": report["Consecutive Absentees"],
                    "5 Absents (at least 10 sessions)": report["Five Absent"],
                    "10 Absents (at least 25 sessions)": report["Ten Absent"],
                    "Discontinued": report["Discontinued"]
                }
    
                # Ensure the lists are padded to the same length
                padded_data = pad_lists(list(summary_data.values()))
    
                # Create a DataFrame for the summary data
                summary_df = pd.DataFrame(
                    padded_data, 
                    index=summary_data.keys()
                ).transpose()
                
                summary_df.to_excel(writer, sheet_name=f"{class_name[:20]} Summary", index=False, startrow=1)
    
                # Collect trainer report data
                trainer_report_data.append({
                    "Class": class_name,
                    "Trainer": report["Trainer"],
                    "Last Updated Date": report["Last date"]
                })
    
            # Add the Trainer's Report sheet
            trainer_report_df = pd.DataFrame(trainer_report_data)
            trainer_report_df.to_excel(writer, sheet_name="Trainer's Report", index=False)
    
        # Get the workbook object from the writer
        workbook = writer.book
        
        for class_name in reports.keys():
            for sheet_name_suffix in [" Data", " Summary"]:
                sheet_name = f"{class_name[:20]}{sheet_name_suffix}"
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Insert the class name in the first row and merge cells
                    sheet.insert_rows(1)
                    sheet["A1"] = class_name
                    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
                    
                    # Align the text to the left
                    sheet["A1"].alignment = Alignment(horizontal="left")
    
        # Save the modified workbook back to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    def find_trainer_notes(excel_file):
        # Load the workbook
        wb = load_workbook(excel_file, read_only=True)
        
        # Check sheet names and find "Teachers Note"
        sheet_names = wb.sheetnames
        if "Teachers Note" not in sheet_names:
            return None
        
        # Load the sheet
        ws = wb["Teachers Note"]
        
        # Search for "Trainer: " in the sheet
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and cell.startswith("Trainer: "):
                    return cell.split("Trainer: ")[-1].strip()
        
        return None
    
    # Function to parse uploaded files
    def parse_uploaded_files(uploaded_files):
        reports = {}
        for name, file_obj in uploaded_files:
            try:
                df = pd.read_excel(file_obj)
                class_name = name.split('.')[0]
                reports[class_name] = {
                    "Attendance Data": calculate_attendance(df),
                    "Low Attendance": students_left(df),
                    "Consecutive Absentees": find_consecutive_absentees(df),
                    "Five Absent": find_absentees(df)[0],
                    "Ten Absent": find_absentees(df)[1],
                    "Discontinued": find_trainer_notes(file_obj),
                    "Trainer": "TBA",
                    "Last date": "24-09-23"
                }
            except Exception as e:
                st.warning(f"Failed to parse {name}: {e}")
    
        return reports
    
    # Sidebar
    st.sidebar.title("File Upload")
    uploaded_files = st.sidebar.file_uploader("Upload your attendance files (Excel format)", type=["xlsx"], accept_multiple_files=True)
    
    # Fetch sample files from GitHub (for demonstration purposes)
    if st.sidebar.button("Use Sample Files"):
        uploaded_files = fetch_sample_files()
    
    if not uploaded_files:
        st.info("Please upload your attendance files or use sample files.")
        st.stop()
    
    # Process uploaded files
    reports = parse_uploaded_files(uploaded_files)
    
    # Generate Excel report
    if st.button("Generate Excel Report"):
        with st.spinner('Generating Excel report...'):
            excel_data = generate_excel(reports)
            st.markdown(get_binary_file_downloader_html(excel_data, 'Excel Report'), unsafe_allow_html=True)
    
    # Display the attendance data
    for class_name, report in reports.items():
        st.subheader(f"{class_name} Attendance Report")
        st.dataframe(report["Attendance Data"].style.apply(highlight_rows, condition_list=report["Low Attendance"]), height=600)
    
    # Display additional insights
    st.subheader("Additional Insights")
    
    st.markdown("**Students who left:**")
    left_students = []
    for class_name, report in reports.items():
        left_students.extend(report["Low Attendance"])
    left_students = list(set(left_students))  # Remove duplicates
    st.write(left_students if left_students else "No students found.")
    
    st.markdown("**Students with 3 consecutive absences:**")
    consecutive_absentees = []
    for class_name, report in reports.items():
        consecutive_absentees.extend(report["Consecutive Absentees"])
    consecutive_absentees = list(set(consecutive_absentees))  # Remove duplicates
    st.write(consecutive_absentees if consecutive_absentees else "No students found.")
    
    st.markdown("**Students with 5 or more absences:**")
    five_absentees = []
    for class_name, report in reports.items():
        five_absentees.extend(report["Five Absent"])
    five_absentees = list(set(five_absentees))  # Remove duplicates
    st.write(five_absentees if five_absentees else "No students found.")
    
    st.markdown("**Students with 10 or more absences:**")
    ten_absentees = []
    for class_name, report in reports.items():
        ten_absentees.extend(report["Ten Absent"])
    ten_absentees = list(set(ten_absentees))  # Remove duplicates
    st.write(ten_absentees if ten_absentees else "No students found.")
    
    st.markdown("**Trainer's Notes:**")
    trainer_notes = []
    for class_name, report in reports.items():
        if report["Discontinued"]:
            trainer_notes.append(f"{class_name}: {report['Discontinued']}")
    st.write(trainer_notes if trainer_notes else "No notes found.")
    
    
if __name__ == "__main__":
    run_main_app()
