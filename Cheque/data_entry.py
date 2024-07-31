from tkinter import Tk, Label, Entry, Button, Frame, filedialog, W, E
from PIL import Image, ImageTk
import os
import openpyxl
import traceback

# Function to save data to Excel file
def save_to_excel(excel_filename, filename, cheque_number, amount, account_number, name):
    try:
        if not os.path.exists(excel_filename):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["File name", "Cheque Number", "Amount", "Account Number", "Name"])
            wb.save(excel_filename)
        else:
            wb = openpyxl.load_workbook(excel_filename)

        sheet = wb.active
        found = False
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == filename:
                # If duplicate found, update existing row with new data
                row_index = row[0].row
                sheet.cell(row=row_index, column=2).value = cheque_number
                sheet.cell(row=row_index, column=3).value = amount
                sheet.cell(row=row_index, column=4).value = account_number
                sheet.cell(row=row_index, column=5).value = name
                found = True
                break

        # If no duplicate found and fields are not empty, append new data
        if not found and any([cheque_number, amount, account_number, name]):
            sheet.append([filename, cheque_number, amount, account_number, name])

        wb.save(excel_filename)
    except PermissionError as e:
        print(f"PermissionError: {e}")
        traceback.print_exc()
    except Exception as e:
        print(f"Error occurred while saving to Excel: {e}")
        traceback.print_exc()

class ChequeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Cheque Viewer and Data Entry")
        self.root.geometry("1500x600")  # Set initial window size

        self.current_index = 0
        self.images = self.load_images()

        # Main frame
        self.main_frame = Frame(root)
        self.main_frame.pack(expand=True, fill='both')

        self.filename_var = Label(self.main_frame, text="", font=('Arial', 14))
        self.filename_var.grid(row=0, column=0, sticky=W, pady=15, padx=30)

        self.file_count_label = Label(self.main_frame, text="", font=('Arial', 14))
        self.file_count_label.grid(row=0, column=1, sticky=E, padx=30)

        self.index_entry = Entry(self.main_frame, font=('Arial', 14), width=5)
        self.index_entry.grid(row=0, column=1, sticky=W, padx=15)
        self.index_entry.bind('<Return>', self.go_to_index)

        self.image_label = Label(self.main_frame)
        self.image_label.grid(row=1, column=0, columnspan=2, padx=20, pady=20)

        # Data Frame
        self.data_frame = Frame(self.main_frame, pady=20)
        self.data_frame.grid(row=1, column=2, rowspan=4, sticky='nsew')

        # Labels and Entries
        self.cheque_number_label = Label(self.data_frame, text="Cheque Number:", font=('Arial', 14))
        self.cheque_number_label.grid(row=1, column=0, sticky=W, pady=15, padx=15)

        self.cheque_number_entry = Entry(self.data_frame, state='normal', font=('Arial', 16), width=25)
        self.cheque_number_entry.grid(row=2, column=0, sticky=W, padx=15, ipady=3)

        self.amount_label = Label(self.data_frame, text="Amount:", font=('Arial', 14))
        self.amount_label.grid(row=3, column=0, sticky=W, pady=15, padx=15)

        self.amount_entry = Entry(self.data_frame, state='normal', font=('Arial', 16), width=25)
        self.amount_entry.grid(row=4, column=0, sticky=W, padx=15, ipady=3)

        self.account_number_label = Label(self.data_frame, text="Account Number:", font=('Arial', 14))
        self.account_number_label.grid(row=5, column=0, sticky=W, pady=15, padx=15)

        self.account_number_entry = Entry(self.data_frame, state='normal', font=('Arial', 16), width=25)
        self.account_number_entry.grid(row=6, column=0, sticky=W, padx=15, ipady=3)

        self.name_label = Label(self.data_frame, text="Name:", font=('Arial', 14))
        self.name_label.grid(row=7, column=0, sticky=W, pady=15, padx=15)

        self.name_entry = Entry(self.data_frame, state='normal', font=('Arial', 16), width=25)
        self.name_entry.grid(row=8, column=0, sticky=W, padx=15, ipady=3)

        # Buttons with fixed width and aligned to left and right
        self.next_button = Button(self.data_frame, text="Next", command=self.show_next, font=('Arial', 14), width=10)
        self.next_button.grid(row=9, column=0, pady=15, sticky=E)

        self.previous_button = Button(self.data_frame, text="Previous", command=self.show_previous, font=('Arial', 14), width=10)
        self.previous_button.grid(row=9, column=0, pady=15, sticky=W)

        # Dictionary to store modified data for each filename
        self.data_modified = {}

        self.update_display()

        # Force update of the window's layout
        self.root.update_idletasks()
        self.root.update()

        # Bring the data frame and entry widgets to the foreground
        self.data_frame.lift()
        self.cheque_number_entry.lift()
        self.amount_entry.lift()
        self.account_number_entry.lift()
        self.name_entry.lift()

        # Ensure the root window has focus
        self.root.focus_force()

        # Set cheque number entry on auto focus
        self.cheque_number_entry.focus_set()

        # Bind Enter key to move focus to next widget
        self.cheque_number_entry.bind('<Return>', self.focus_next_widget)
        self.amount_entry.bind('<Return>', self.focus_next_widget)
        self.account_number_entry.bind('<Return>', self.focus_next_widget)
        self.name_entry.bind('<Return>', self.show_next)

    def focus_next_widget(self, event):
        event.widget.tk_focusNext().focus()

    def load_images(self):
        folder_path = filedialog.askdirectory(title="Select Folder with Images")
        if not folder_path:
            exit()

        self.folder_name = os.path.basename(folder_path)
        self.excel_filename = f"{self.folder_name}.xlsx"
        image_files = [file for file in os.listdir(folder_path) if file.lower().endswith(('f.tif', 'f.jpg', 'f.jpeg'))]

        images = []
        for file in image_files:
            image_path = os.path.join(folder_path, file)
            image = Image.open(image_path)
            images.append((image, file))

        # Load Excel data to skip already entered files
        self.existing_files = self.load_excel_data()

        # Find the first new image
        for index, (_, filename) in enumerate(images):
            if filename not in self.existing_files:
                self.current_index = index
                break

        return images

    def update_display(self):
        if self.current_index >= len(self.images):
            self.current_index = 0
        elif self.current_index < 0:
            self.current_index = len(self.images) - 1

        image, filename = self.images[self.current_index]
        self.filename_var.config(text=filename)
        self.file_count_label.config(text=f"{self.current_index + 1}/{len(self.images)}")
        self.index_entry.delete(0, 'end')
        self.index_entry.insert(0, str(self.current_index + 1))

        image, filename = self.images[self.current_index]
        self.filename_var.config(text=filename)

        # Check if filename is already in Excel
        self.load_image_data(filename)

        # Resize image to fit label
        image.thumbnail((1000, 600))  # Adjust size as needed
        self.image_tk = ImageTk.PhotoImage(image)
        self.image_label.config(image=self.image_tk)

    def load_excel_data(self):
        existing_files = set()
        try:
            if os.path.exists(self.excel_filename):
                wb = openpyxl.load_workbook(self.excel_filename)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    existing_files.add(row[0])
        except Exception as e:
            print(f"Error loading data from Excel: {e}")
            traceback.print_exc()
        return existing_files

    def load_image_data(self, filename):
        try:
            if os.path.exists(self.excel_filename):
                wb = openpyxl.load_workbook(self.excel_filename)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] == filename:
                        self.cheque_number_entry.delete(0, 'end')
                        self.cheque_number_entry.insert(0, row[1] if row[1] else '')
                        self.amount_entry.delete(0, 'end')
                        self.amount_entry.insert(0, row[2] if row[2] else '')
                        self.account_number_entry.delete(0, 'end')
                        self.account_number_entry.insert(0, row[3] if row[3] else '')
                        self.name_entry.delete(0, 'end')
                        self.name_entry.insert(0, row[4] if row[4] else '')
                        return
                # If not found, clear the entries
                self.cheque_number_entry.delete(0, 'end')
                self.amount_entry.delete(0, 'end')
                self.account_number_entry.delete(0, 'end')
                self.name_entry.delete(0, 'end')
        except Exception as e:
            print(f"Error loading data from Excel: {e}")
            traceback.print_exc()

    def save_current_data(self):
        excel_filename = f"{self.folder_name}.xlsx"
        filename = self.filename_var.cget("text")
        cheque_number = self.cheque_number_entry.get()
        amount = self.amount_entry.get()
        account_number = self.account_number_entry.get()
        name = self.name_entry.get()

        # Save current data into data_modified dictionary
        self.data_modified[filename] = {
            "cheque_number": cheque_number,
            "amount": amount,
            "account_number": account_number,
            "name": name
        }

        # Save data to Excel
        save_to_excel(excel_filename, filename, cheque_number, amount, account_number, name)

    def show_previous(self):
        self.save_current_data()
        self.current_index -= 1
        self.update_display()

        # Set cheque number entry on auto focus
        self.cheque_number_entry.focus_set()

    def show_next(self, event=None):
        self.save_current_data()
        self.current_index += 1
        self.update_display()

        # Set cheque number entry on auto focus
        self.cheque_number_entry.focus_set()

    def go_to_index(self, event=None):
        try:
            index = int(self.index_entry.get()) - 1
            if 0 <= index < len(self.images):
                self.save_current_data()
                self.current_index = index
                self.update_display()
                # Set cheque number entry on auto focus
                self.cheque_number_entry.focus_set()
            else:
                raise ValueError("Index out of range")
        except ValueError as e:
            print(f"Invalid index: {e}")

if __name__ == "__main__":
    root = Tk()
    app = ChequeApp(root)
    root.mainloop()
